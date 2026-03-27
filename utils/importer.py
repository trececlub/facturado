from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


FIELD_ALIASES = {
    "nombre": ["nombre", "name", "first_name"],
    "apellido": ["apellido", "segundo_nombre", "last_name", "surname"],
    "cargo": ["cargo", "categoria", "role", "position"],
    "id_interno": ["id_interno", "id", "employee_id", "internal_id"],
    "fecha": ["fecha", "date"],
    "sede": ["sede", "direccion", "site", "address"],
    "observaciones": ["observaciones", "notes", "remarks"],
    "foto_path": ["foto", "imagen", "photo", "foto_path", "image_path"],
    "barcode": ["barcode", "codigo_barra", "codigo"],
    "qr": ["qr", "qr_text", "qrcode"]
}


def _normalize_key(key: str) -> str:
    k = key.strip().lower()
    for canonical, aliases in FIELD_ALIASES.items():
        if k in aliases:
            return canonical
    return k


def _normalize_row(row: dict) -> dict:
    normalized = {}
    for k, v in row.items():
        canonical = _normalize_key(str(k))
        normalized[canonical] = "" if v is None else str(v)
    return normalized


def import_csv(path: str | Path) -> list[dict]:
    rows: list[dict] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(_normalize_row(row))
    return rows


def import_excel(path: str | Path) -> list[dict]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    header: list[str] = []
    rows: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            continue

        as_dict = {header[idx]: row[idx] if idx < len(row) else "" for idx in range(len(header))}
        rows.append(_normalize_row(as_dict))

    wb.close()
    return rows


def import_records(path: str | Path) -> list[dict]:
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".csv":
        return import_csv(p)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return import_excel(p)

    raise ValueError("Formato no soportado. Usa CSV o Excel (.xlsx).")
